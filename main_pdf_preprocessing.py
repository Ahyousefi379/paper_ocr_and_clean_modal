import os
from pathlib import Path
from typing import List, Dict
import time

from openpyxl import Workbook, load_workbook

from pdf_references_cleaner import remove_pages_after_references
from merger_ocrs_util import symmetrical_merge
from validate_pdf import validate_pdfs
from olmOCR_api_util import OlmOcrPollingClient
#from Olmocr_modal_api_util import OlmOcrPollingClient

import pandas as pd
from headings_fix_util import ScientificTextToMarkdownConverter


# --- CONFIGURATION ---
BASE_DIR = Path(r"H:\python_projects\scientific\PDT_DATA_FILES\papers\gigalib_papers")
RAW_PDFS_DIR = BASE_DIR / "raw"
CLEANED_PDFS_DIR = BASE_DIR / "cleaned"
RAW_OCR_DIR = BASE_DIR / "raw ocr"
MD_DIR = BASE_DIR / "md"

#SERVER_URL = "https://8000-01k4ymt1036tngva09ywdk72z6.cloudspaces.litng.ai"  #account = ahyousefi379+1
#SERVER_URL = "https://8000-01k63fzfe4jzfjv93prex1pk7v.cloudspaces.litng.ai"  #account = ahyousefi379y??
#SERVER_URL=  "https://8000-01k6516jt155gsnbey7qymq56x.cloudspaces.litng.ai"  #ahyousefi379channel
#SERVER_URL = "https://8000-01k6crhbwkcph31rxs9551d56x.cloudspaces.litng.ai"  #ahyousefi379
#SERVER_URL = "https://8000-01k9m16573j2yfbyadhqghew5c.cloudspaces.litng.ai"  #themadscientist

#SERVER_URL = "https://ahyousefi379--olmocr-api-fastapi-app.modal.run" #modal ahyousefi379
SERVER_URL="https://ahyousefi379y--olmocr-api-fastapi-app.modal.run" #modal ahyousefi379y


POLL_INTERVAL_SECONDS = 5
TOTAL_WAIT_MINUTES = 4
NUM_OCR_VERSIONS = 1
MAX_RETRY_ATTEMPTS = 1


def setup_directories():
    print("Setting up directories...")
    for dir_path in [CLEANED_PDFS_DIR, RAW_OCR_DIR, MD_DIR]:
        dir_path.mkdir(parents=True, exist_ok=True)


def clean_pdfs(source_dir: Path, target_dir: Path):
    print(f"Scanning for PDFs in '{source_dir}'...")
    raw_pdf_files = list(source_dir.glob('*.pdf'))

    if not raw_pdf_files:
        print(f"No PDF files found in '{source_dir}'.")
        return

    print(f"Found {len(raw_pdf_files)} PDFs. Cleaning references...")
    for pdf_path in raw_pdf_files:
        output_path = target_dir / f"_cleaned_{pdf_path.name}"
        print(f"  - Processing {pdf_path.name}")
        remove_pages_after_references(
            input_pdf_path=str(pdf_path),
            output_pdf_path=str(output_path)
        )
    print("PDF cleaning complete.\n")



def is_ocr_successful(output_path: Path) -> bool:
    if not output_path.exists():
        return False
    try:
        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            return len(content) > 100
    except Exception:
        return False


def get_base_name_from_cleaned(pdf_path: Path) -> str:
    return pdf_path.stem.replace('_cleaned_', '')


def perform_ocr_with_validation(source_files: List[Path], target_dir: Path, save_reports_dir:Path):
    print("Starting OCR process...")
    if not source_files:
        print("No valid cleaned PDFs to process.")
        return

    client = OlmOcrPollingClient(
        base_url=SERVER_URL,
        json_report_file=str(save_reports_dir / "ocr_report.json")
    )

    successful_pdfs = []

    # Excel path to record failed PDFs
    excel_path = save_reports_dir / "failed_ocr_pdfs.xlsx"
    
    # Initialize Excel if it doesn't exist
    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Failed OCR PDFs"
        ws.append(["Filename", "Reason"])
        wb.save(str(excel_path))

    for pdf_path in source_files:
        base_name = get_base_name_from_cleaned(pdf_path)
        cleaned_filename = pdf_path.name

        print(f"  - Processing '{cleaned_filename}' (need {NUM_OCR_VERSIONS} successful versions)...")

        existing_successes = client.get_existing_success_count(cleaned_filename)
        print(f"    - Found {existing_successes} existing successful attempts")

        successful_versions = existing_successes
        failed_reason = ""

        for version in range(existing_successes + 1, NUM_OCR_VERSIONS + 1):
            output_path = target_dir / f"{base_name}_attempt_{version}.md"

            print(f"    - Working on version {version}/{NUM_OCR_VERSIONS}")

            version_successful = False
            for attempt in range(1, MAX_RETRY_ATTEMPTS + 1):
                print(f"      - Retry {attempt}: Processing version {version}")

                try:
                    success = client.process_document(
                        file_path=str(pdf_path),
                        save_output=True,
                        output_dir=str(output_path.parent),
                        output_filename=output_path.name,
                        poll_interval_sec=POLL_INTERVAL_SECONDS,
                        total_wait_min=TOTAL_WAIT_MINUTES,
                        attempt_id=f"attempt_{version}",
                        force_process=True
                    )

                    if success and is_ocr_successful(output_path):
                        print(f"      - Version {version} completed successfully")
                        version_successful = True
                        successful_versions += 1
                        break
                    else:
                        failed_reason = "No valid OCR output"
                        print(f"      - Version {version} failed (no valid output)")

                except Exception as e:
                    failed_reason = str(e)
                    print(f"      - Version {version} failed with error: {failed_reason}")

                if attempt < MAX_RETRY_ATTEMPTS:
                    print(f"      - Retrying version {version} in 10 seconds...")
                    time.sleep(10)

            if not version_successful:
                print(f"    - Version {version} failed after {MAX_RETRY_ATTEMPTS} attempts")

        if successful_versions >= NUM_OCR_VERSIONS:
            print(f"  ✓ '{cleaned_filename}' completed successfully ({successful_versions}/{NUM_OCR_VERSIONS} versions)")
            successful_pdfs.append(cleaned_filename)
        else:
            print(f"  ✗ '{cleaned_filename}' failed - only {successful_versions}/{NUM_OCR_VERSIONS} versions succeeded")
            
            # --- Record failed PDF in Excel immediately ---
            try:
                wb = load_workbook(str(excel_path))
                ws = wb.active
                ws.append([cleaned_filename, failed_reason])
                wb.save(str(excel_path))
                print(f"    - Recorded failed PDF in Excel: {excel_path.name}")
            except Exception as e:
                print(f"    - Could not write failed PDF to Excel: {e}")

    print(f"\nOCR processing complete!")
    print(f"Successful PDFs: {len(successful_pdfs)}")
    
    # Optionally, print a summary of failed PDFs from Excel
    if excel_path.exists():
        print(f"Failed PDFs recorded in Excel: {excel_path}")
    print()


def post_process_ocr(source_dir: Path, target_dir: Path, excel_filepath: Path):
    """
    Post-process OCR markdown files by adding abstracts from Excel and marking as processed.
    
    Args:
        source_dir: Directory containing OCR attempt markdown files (*_attempt*.md)
        target_dir: Directory where final processed files will be saved
        excel_filepath: Path to the Excel file containing paper information
    """
    print("Starting OCR post-processing...")
    
    # Track missing matches for final report
    missing_matches = []
    
    # Create log file path
    log_file = target_dir / "matching_debug_log.txt"
    
    # Create section reports directory
    reports_dir = target_dir / "section_reports"
    reports_dir.mkdir(exist_ok=True, parents=True)
    
    # Clear existing log file
    if log_file.exists():
        log_file.unlink()
    
    # Write log header
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("OCR POST-PROCESSING MATCHING LOG\n")
        f.write("="*80 + "\n")
        f.write(f"Started at: {pd.Timestamp.now()}\n")
        f.write(f"Source directory: {source_dir}\n")
        f.write(f"Target directory: {target_dir}\n")
        f.write(f"Excel file: {excel_filepath}\n")
        f.write("="*80 + "\n\n")
    
    # Find all OCR attempt files
    attempt_files: Dict[str, List[Path]] = {}
    for md_file in source_dir.glob("*_attempt*.md"):
        parts = md_file.stem.split("_attempt")
        if len(parts) == 2:
            base_name = parts[0]
            attempt_files.setdefault(base_name, []).append(md_file)

    if not attempt_files:
        print(f"No OCR attempt markdown files found in '{source_dir}'.")
        return

    # Create target directory if it doesn't exist
    target_dir.mkdir(exist_ok=True, parents=True)
    
    # Process each base file
    for base_name, md_files in attempt_files.items():
        print(f"\n  Processing '{base_name}'...")
        
        # Use the first attempt file (or you could merge multiple attempts)
        input_file = md_files[0]
        if len(md_files) > 1:
            print(f"    Note: Found {len(md_files)} attempts, using first one: {input_file.name}")
        
        output_path = target_dir / f"_final_{base_name}.md"
        
        try:
            # Create converter with the input file, Excel database, and log file
            converter = ScientificTextToMarkdownConverter(input_file, excel_filepath, log_file)
            
            # Try to save the file - util will only create it if match was found
            if not converter.save_to_file(output_path):
                # File was not created because no match was found
                missing_matches.append(base_name)
            
        except Exception as e:
            print(f"    ✗ Error processing '{base_name}': {e}")
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(f"    ✗ ERROR: {e}\n\n")
            missing_matches.append(base_name)
    
    # Generate missing matches report
    print("\n" + "="*60)
    print("POST-PROCESSING COMPLETE")
    print("="*60)
    
    if missing_matches:
        print(f"\n⚠ {len(missing_matches)} file(s) failed to match with Excel titles:\n")
        
        # Create report file
        report_file = target_dir / "missing_matches_report.txt"
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write("MISSING FILENAME-TITLE MATCHES REPORT\n")
            f.write("="*60 + "\n\n")
            f.write(f"Total files that failed to match: {len(missing_matches)}\n\n")
            f.write("Base filenames (without _attempt suffix):\n")
            f.write("-"*60 + "\n")
            for filename in missing_matches:
                f.write(f"  - {filename}\n")
            f.write("\n" + "="*60 + "\n")
            f.write("Please check these filenames and update the Excel 'Title' column\n")
            f.write("or rename the files to match existing titles.\n")
        
        print(f"✓ Missing matches report saved to: {report_file}")
        print("\nMissing files:")
        for filename in missing_matches:
            print(f"  - {filename}")
    else:
        print("\n✓ All files successfully matched and processed!")
    
    print(f"\nTotal processed: {len(attempt_files)}")
    print(f"Successful: {len(attempt_files) - len(missing_matches)}")
    print(f"Failed: {len(missing_matches)}")
    
    # Write log footer
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write("\n" + "="*80 + "\n")
        f.write("SUMMARY\n")
        f.write("="*80 + "\n")
        f.write(f"Total files: {len(attempt_files)}\n")
        f.write(f"Successful matches: {len(attempt_files) - len(missing_matches)}\n")
        f.write(f"Failed matches: {len(missing_matches)}\n")
        f.write(f"Completed at: {pd.Timestamp.now()}\n")
        f.write("="*80 + "\n")
    
    print(f"\n✓ Detailed matching log saved to: {log_file}")
    print("\nPost-processing complete.\n")



def main():
    setup_directories()

    # Step 1: Remove reference pages
    #clean_pdfs(source_dir=RAW_PDFS_DIR,
    #           target_dir=CLEANED_PDFS_DIR)
 
    # Step 2: Validate PDFs
    valid_pdfs = validate_pdfs(source_dir=CLEANED_PDFS_DIR,
                               save_report_dir=BASE_DIR)
#
    # Step 3: Run OCR only on valid PDFs
    perform_ocr_with_validation(source_files=valid_pdfs,
                                target_dir=RAW_OCR_DIR,
                                save_reports_dir=BASE_DIR)

    # Step 4: Merge OCR results
    #post_process_ocr(source_dir=RAW_OCR_DIR, target_dir=MD_DIR,excel_filepath= "H://python_projects//scientific//pdt_paper_filter//#papers_list.xlsx")

    print("Pipeline finished.")


if __name__ == "__main__":
    main()
